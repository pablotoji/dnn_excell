# -*- coding: utf-8 -*-
"""
Created on Wed Mar  5 19:03:38 2025

@author: pablo
"""

import numpy as np
from tensorflow.keras.datasets import mnist
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, InputLayer
from tensorflow.image import resize
from openpyxl import Workbook

# ------------------------------
# Step 1: Load and preprocess MNIST data
# ------------------------------
(x_train, y_train), (x_test, y_test) = mnist.load_data()
x_train = x_train.astype('float32') / 255.0

def preprocess_image(img, new_size=(10, 10), threshold=0.5):
    """
    Resize a 28x28 image to new_size and binarize it.
    """
    img_resized = resize(img[..., np.newaxis], new_size, method='bilinear').numpy().squeeze()
    img_binary = (img_resized > threshold).astype(int)
    return img_binary

num_samples = 5000  # adjust as needed
x_train_small = np.array([preprocess_image(img) for img in x_train[:num_samples]])
y_train_small = y_train[:num_samples]
x_train_flat = x_train_small.reshape(num_samples, -1)  # shape: (num_samples, 100)

# ------------------------------
# Step 2: Define and train the DNN model
# ------------------------------
model = Sequential([
    InputLayer(input_shape=(100,)),
    Dense(64, activation='relu'),
    Dense(10, activation='softmax')
])
model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])
model.fit(x_train_flat, y_train_small, epochs=10, batch_size=32, validation_split=0.1)

# ------------------------------
# Step 3: Extract weights and biases
# ------------------------------
# weights_biases[0]: weights for Dense layer 1, shape (100, 64)
# weights_biases[1]: biases for Dense layer 1, shape (64,)
# weights_biases[2]: weights for Dense layer 2, shape (64, 10)
# weights_biases[3]: biases for Dense layer 2, shape (10,)
weights_biases = model.get_weights()
w1, b1, w2, b2 = weights_biases

# ------------------------------
# Step 4: Create the Excel workbook with complete formulas
# ------------------------------
wb = Workbook()

# ----- Sheet1: Input grid, flattened input, and predicted output -----
ws_input = wb.active
ws_input.title = "Sheet1"

# Label for the input grid
ws_input["A1"] = "Input 10x10 Grid (enter 'x' for a stroke)"

# Create a 10x10 input grid in B2:K11 (user enters "x" for stroke)
for i in range(10):
    for j in range(10):
        ws_input.cell(row=2 + i, column=2 + j, value="")  # empty cell

# Create a helper area to flatten the 10x10 grid.
# We place the flattened vector in column M (cells M2:M101).
# The formula converts each corresponding cell in B2:K11: if it equals "x", returns 1, else 0.
flatten_formula = '=IF(INDEX($B$2:$K$11, INT((ROW()-2)/10)+1, MOD(ROW()-2,10)+1)="x", 1, 0)'
for i in range(100):
    ws_input.cell(row=2 + i, column=13, value=flatten_formula)  # column 13 is "M"

# Create a cell for predicted output.
# Here we assume that the OutputLayer sheet (to be created) will have the 10 pre-activation outputs in A2:A11.
# The formula picks the index (digit) corresponding to the maximum output.
# We use CHOOSE combined with MATCH.
pred_formula = '=CHOOSE(MATCH(MAX(OutputLayer!$A$2:$A$11), OutputLayer!$A$2:$A$11, 0), 0,1,2,3,4,5,6,7,8,9)'
ws_input.cell(row=2, column=15, value=pred_formula)  # cell O2 (column 15)

# ----- Sheets for model parameters -----
# We assume the following ranges:
# Weights1: shape (100,64) will be stored in range A1:BL100 (BL is the 64th column)
# Biases1: shape (64,) in range A1:BL1
# Weights2: shape (64,10) in range A1:J64 (J is the 10th column)
# Biases2: shape (10,) in range A1:J1

ws_w1 = wb.create_sheet(title="Weights1")
ws_b1 = wb.create_sheet(title="Biases1")
ws_w2 = wb.create_sheet(title="Weights2")
ws_b2 = wb.create_sheet(title="Biases2")

# Write weights for the first Dense layer
for i in range(w1.shape[0]):  # 100 rows
    for j in range(w1.shape[1]):  # 64 columns
        ws_w1.cell(row=i+1, column=j+1, value=float(w1[i, j]))

# Write biases for the first Dense layer (in one row)
for j in range(b1.shape[0]):
    ws_b1.cell(row=1, column=j+1, value=float(b1[j]))

# Write weights for the second Dense layer
for i in range(w2.shape[0]):  # 64 rows
    for j in range(w2.shape[1]):  # 10 columns
        ws_w2.cell(row=i+1, column=j+1, value=float(w2[i, j]))

# Write biases for the second Dense layer (in one row)
for j in range(b2.shape[0]):
    ws_b2.cell(row=1, column=j+1, value=float(b2[j]))

# ----- HiddenLayer sheet: Compute hidden layer activations (ReLU applied) -----
ws_hidden = wb.create_sheet(title="HiddenLayer")
# For each of the 64 hidden neurons, compute:
# =MAX(0, SUMPRODUCT(Sheet1!$M$2:$M$101, INDEX(Weights1!$A$1:$BL$100,0, neuron_index)) + INDEX(Biases1!$A$1:$BL$1,1, neuron_index))
for neuron in range(64):
    # neuron index is (neuron+1)
    formula = f'=MAX(0, SUMPRODUCT(Sheet1!$M$2:$M$101, INDEX(Weights1!$A$1:$BL$100, 0, {neuron+1})) + INDEX(Biases1!$A$1:$BL$1, 1, {neuron+1}))'
    ws_hidden.cell(row=2 + neuron, column=1, value=formula)
# Optionally, add a label in A1
ws_hidden["A1"] = "Hidden Layer Outputs (64 neurons)"

# ----- OutputLayer sheet: Compute output layer pre-activations -----
ws_output = wb.create_sheet(title="OutputLayer")
# For each of the 10 output neurons, compute:
# =SUMPRODUCT(HiddenLayer!$A$2:$A$65, INDEX(Weights2!$A$1:$J64, 0, neuron_index)) + INDEX(Biases2!$A$1:$J1, 1, neuron_index)
for neuron in range(10):
    formula = f'=SUMPRODUCT(HiddenLayer!$A$2:$A$65, INDEX(Weights2!$A$1:$J64, 0, {neuron+1})) + INDEX(Biases2!$A$1:$J1, 1, {neuron+1})'
    ws_output.cell(row=2 + neuron, column=1, value=formula)
ws_output["A1"] = "Output Layer Pre-activations (10 neurons)"

# ----- Save the complete Excel workbook -----
excel_filename = "complete_dnn_model.xlsx"
wb.save(excel_filename)
print(f"Excel file '{excel_filename}' has been created with the complete forward-pass formulas.")
